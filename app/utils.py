from typing import Dict, Any, Optional
import requests
import json
from datetime import datetime, timedelta
from io import BytesIO
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.chart import LineChart, Reference
import os

# Cache dla kursu EUR (ważny przez 1 dzień)
_eur_rate_cache: Dict[str, Any] = {"rate": None, "timestamp": None}

def get_eur_rate_from_nbp() -> float:
    """
    Pobiera aktualny kurs EUR/PLN z API Narodowego Banku Polskiego.
    Używa cache'u aby ograniczyć liczbę zapytań do API.
    
    Returns:
        float: Kurs EUR/PLN
    """
    global _eur_rate_cache
    
    # Sprawdź czy mamy świeży kurs w cache (max 1 dzień)
    if (_eur_rate_cache["rate"] is not None and 
        _eur_rate_cache["timestamp"] is not None):
        
        time_diff = datetime.now() - _eur_rate_cache["timestamp"]
        if time_diff.total_seconds() < 24 * 3600:  # 24 godziny
            return _eur_rate_cache["rate"]
    
    try:
        # Pobierz aktualny kurs EUR z NBP API
        response = requests.get(
            "https://api.nbp.pl/api/exchangerates/rates/a/eur/",
            headers={"Accept": "application/json"},
            timeout=10
        )
        
        if response.status_code == 200:
            data = response.json()
            rate = data["rates"][0]["mid"]
            
            # Zapisz w cache
            _eur_rate_cache["rate"] = rate
            _eur_rate_cache["timestamp"] = datetime.now()
            
            return rate
        else:
            # Fallback rate jeśli API nie działa
            return 4.30  # Przybliżony kurs EUR/PLN
            
    except Exception as e:
        print(f"Błąd przy pobieraniu kursu EUR: {e}")
        # Fallback rate
        return 4.30

def convert_pln_to_eur(amount_pln: float) -> float:
    """
    Przelicza kwotę z PLN na EUR używając kursu NBP.
    
    Args:
        amount_pln (float): Kwota w PLN
    
    Returns:
        float: Kwota w EUR
    """
    eur_rate = get_eur_rate_from_nbp()
    return round(amount_pln / eur_rate, 2)

def convert_eur_to_pln(amount_eur: float) -> float:
    """
    Przelicza kwotę z EUR na PLN używając kursu NBP.
    
    Args:
        amount_eur (float): Kwota w EUR
    
    Returns:
        float: Kwota w PLN
    """
    eur_rate = get_eur_rate_from_nbp()
    return round(amount_eur * eur_rate, 2)

def calculate_acos(sales: float, spend: float, margin: float) -> Dict[str, Any]:
    """
    Oblicza wskaźniki ACOS, ROI, zysk i sprawdza rentowność kampanii.
    
    Args:
        sales (float): Prognozowana sprzedaż
        spend (float): Prognozowane wydatki na reklamę
        margin (float): Marża brutto w procentach
    
    Returns:
        Dict[str, Any]: Słownik z obliczonymi wskaźnikami
    """
    
    # Obliczenie ACOS (Advertising Cost of Sales)
    acos = (spend / sales) * 100 if sales > 0 else 0
    
    # Obliczenie ROI (Return on Investment)
    roi = ((sales - spend) / spend) * 100 if spend > 0 else 0
    
    # Poprawione obliczenie zysku
    # Zysk brutto ze sprzedaży
    gross_profit = sales * (margin / 100)
    # Zysk całkowity po odjęciu kosztów reklamy
    total_profit = gross_profit - spend
    # Profit per sale - zakładając że każda sprzedaż to jedno zamówienie
    profit_per_sale = total_profit  # Dla podstawowego kalkulatora
    
    # Break-even ACOS (punkt rentowności)
    break_even_acos = margin
    
    # Sprawdzenie rentowności
    is_profitable = acos <= margin if margin > 0 else False
    
    # Przygotowanie komunikatów
    profitability_message = ""
    if not is_profitable and acos > 0:
        profitability_message = f"⚠️ UWAGA: Kampania jest nierentowna! ACOS ({acos:.2f}%) przekracza marżę ({margin:.2f}%)"
    elif is_profitable and acos > 0:
        profitability_message = f"✅ Kampania jest rentowna! ACOS ({acos:.2f}%) mieści się w marży ({margin:.2f}%)"
    
    # Określenie statusu rentowności
    profitability_status = "profitable" if is_profitable else "unprofitable"
    
    return {
        "acos": round(acos, 2),
        "roi": round(roi, 2),
        "profit": round(total_profit, 2),
        "profit_per_sale": round(profit_per_sale, 2),
        "break_even_acos": round(break_even_acos, 2),
        "is_profitable": is_profitable,
        "profitability_message": profitability_message,
        "profitability_status": profitability_status,
        "sales": sales,
        "spend": spend,
        "margin": margin
    }

def calculate_forecast_from_metrics(
    gross_margin: float,
    target_aov: float,
    target_ctr: float,
    target_cpc: float,
    target_cvr: float,
    impressions: int,
    currency: str = "EUR"
) -> Dict[str, Any]:
    """
    Oblicza prognozę kampanii na podstawie zaawansowanych metryk w EUR.
    Używa poprawionej formuły ACOS: (Bid × Clicks) / [(Impressions × CTR%) × CVR% × AOV]
    
    Args:
        gross_margin (float): Marża brutto w procentach
        target_aov (float): Docelowa wartość średniego zamówienia w EUR
        target_ctr (float): Docelowy CTR w procentach
        target_cpc (float): Docelowy koszt za kliknięcie w EUR (Bid)
        target_cvr (float): Docelowy współczynnik konwersji w procentach
        impressions (int): Liczba wyświetleń
        currency (str): Waluta do wyświetlania (EUR)
    
    Returns:
        Dict[str, Any]: Słownik z prognozami i wskaźnikami
    """
    
    # Obliczenia podstawowe zgodnie z formułą ACOS
    clicks = impressions * (target_ctr / 100)
    orders = clicks * (target_cvr / 100)  # Zmiana nazwy z conversions na orders
    ad_spend = clicks * target_cpc  # Ad Spend = Bid × Clicks
    ad_sales = orders * target_aov  # Ad Sales = Orders × AOV
    
    # Obliczenie ACOS według wzoru: ACOS = Ad Spend / Ad Sales
    acos = (ad_spend / ad_sales) * 100 if ad_sales > 0 else 0
    
    # ROI = (Revenue - Cost) / Cost
    roi = ((ad_sales - ad_spend) / ad_spend) * 100 if ad_spend > 0 else 0
    
    # ROAS = Revenue / Ad Spend
    roas = ad_sales / ad_spend if ad_spend > 0 else 0
    
    # Poprawione obliczenie zysku
    gross_profit = ad_sales * (gross_margin / 100)  # Zysk brutto ze sprzedaży
    total_profit = gross_profit - ad_spend  # Zysk całkowity po odjęciu kosztów reklamy
    profit_per_sale = total_profit / orders if orders > 0 else 0  # Zysk na sprzedaż
    
    # Break-even ACOS = marża brutto
    break_even_acos = gross_margin
    
    # Sprawdzenie rentowności
    is_profitable = acos <= gross_margin if gross_margin > 0 and acos > 0 else False
    
    # Przygotowanie komunikatów rentowności
    profitability_message = ""
    if acos > 0:
        if not is_profitable:
            profitability_message = (
                f"⚠️ Nierentowna kampania: ACOS {acos:.0f}% przekracza marżę {gross_margin:.0f}%. "
                f"Rozważ poprawę współczynnika konwersji, obniżenie CPC lub zwiększenie marży produktu."
            )
        else:
            profitability_message = f"✅ Rentowna kampania: ACOS {acos:.0f}% mieści się w marży {gross_margin:.0f}%"
    
    # Określenie statusu rentowności
    profitability_status = "profitable" if is_profitable else "unprofitable"
    
    # Pobierz aktualny kurs EUR/PLN dla informacji
    eur_rate = get_eur_rate_from_nbp()
    
    # Przeliczenia na PLN dla wyświetlania
    projected_sales_pln = round(ad_sales * eur_rate, 0)
    projected_spend_pln = round(ad_spend * eur_rate, 0)
    profit_pln = round(total_profit * eur_rate, 0)
    profit_per_sale_pln = round(profit_per_sale * eur_rate, 0)
    target_aov_pln = round(target_aov * eur_rate, 0)
    target_cpc_pln = round(target_cpc * eur_rate, 2)
    
    return {
        # Wskaźniki podstawowe
        "acos": round(acos, 0),  # Zaokrąglenie do całości jak na screenshocie
        "roi": round(roi, 0),
        "profit": round(total_profit, 0),
        "profit_per_sale": round(profit_per_sale, 0),
        "break_even_acos": round(break_even_acos, 0),
        "is_profitable": is_profitable,
        "profitability_message": profitability_message,
        "profitability_status": profitability_status,
        
        # Prognozowane wyniki (Forecast Results)
        "impressions": impressions,
        "clicks": round(clicks, 0),  # Projected Clicks
        "orders": round(orders, 0),  # Projected Orders (zamiast conversions)
        "projected_sales": round(ad_sales, 0),  # Projected Sales
        "projected_spend": round(ad_spend, 0),  # Projected Spend
        
        # Wartości w PLN dla wyświetlania
        "projected_sales_pln": projected_sales_pln,
        "projected_spend_pln": projected_spend_pln,
        "profit_pln": profit_pln,
        "profit_per_sale_pln": profit_per_sale_pln,
        "target_aov_pln": target_aov_pln,
        "target_cpc_pln": target_cpc_pln,
        
        # Parametry wejściowe
        "gross_margin": gross_margin,
        "target_aov": target_aov,
        "target_ctr": target_ctr,
        "target_cpc": target_cpc,
        "target_cvr": target_cvr,
        
        # Dodatkowe wskaźniki
        "cpm": round((ad_spend / impressions) * 1000, 2) if impressions > 0 else 0,
        "cost_per_conversion": round(ad_spend / orders, 2) if orders > 0 else 0,
        "roas": round(roas, 2),
        
        # Informacje o walucie
        "currency": currency,
        "eur_rate": round(eur_rate, 4),
        "currency_info": f"Kurs EUR/PLN: {eur_rate:.4f} (NBP)"
    }

def calculate_budget_from_tacos(
    target_sales: float,
    target_tacos: float,
    gross_margin: float,
    currency: str = "EUR"
) -> Dict[str, Any]:
    """
    Oblicza budżet marketingowy na podstawie zakładanego TACOS (Total Advertising Cost of Sales).
    TACOS = (Wydatki na marketing / Wartość sprzedaży) * 100
    
    Args:
        target_sales (float): Docelowa wartość sprzedaży w EUR
        target_tacos (float): Zakładany TACOS w procentach
        gross_margin (float): Marża brutto w procentach
        currency (str): Waluta do wyświetlania (EUR)
    
    Returns:
        Dict[str, Any]: Słownik z obliczonymi wskaźnikami budżetu
    """
    
    # Obliczenie budżetu marketingowego na podstawie TACOS
    # TACOS = (Marketing Spend / Sales) * 100
    # Marketing Spend = (TACOS * Sales) / 100
    marketing_budget = (target_tacos * target_sales) / 100
    
    # Obliczenie zysku brutto ze sprzedaży
    gross_profit = target_sales * (gross_margin / 100)
    
    # Obliczenie zysku netto po odjęciu kosztów marketingowych
    net_profit = gross_profit - marketing_budget
    
    # Obliczenie ROI
    roi = ((target_sales - marketing_budget) / marketing_budget) * 100 if marketing_budget > 0 else 0
    
    # Sprawdzenie rentowności
    is_profitable = target_tacos <= gross_margin if gross_margin > 0 else False
    
    # Przygotowanie komunikatów rentowności
    profitability_message = ""
    if target_tacos > 0:
        if not is_profitable:
            profitability_message = (
                f"⚠️ UWAGA: Zakładany TACOS {target_tacos:.1f}% przekracza marżę {gross_margin:.1f}%. "
                f"Kampania może być nierentowna przy takim poziomie wydatków marketingowych."
            )
        else:
            profitability_message = f"✅ Rentowna kampania: TACOS {target_tacos:.1f}% mieści się w marży {gross_margin:.1f}%"
    
    # Określenie statusu rentowności
    profitability_status = "profitable" if is_profitable else "unprofitable"
    
    # Pobierz aktualny kurs EUR/PLN dla informacji
    eur_rate = get_eur_rate_from_nbp()
    
    # Przeliczenia na PLN dla wyświetlania
    target_sales_pln = round(target_sales * eur_rate, 0)
    marketing_budget_pln = round(marketing_budget * eur_rate, 0)
    gross_profit_pln = round(gross_profit * eur_rate, 0)
    net_profit_pln = round(net_profit * eur_rate, 0)
    
    return {
        # Wskaźniki podstawowe
        "target_sales": round(target_sales, 0),
        "target_tacos": round(target_tacos, 1),
        "marketing_budget": round(marketing_budget, 0),
        "gross_profit": round(gross_profit, 0),
        "net_profit": round(net_profit, 0),
        "roi": round(roi, 1),
        "gross_margin": gross_margin,
        "is_profitable": is_profitable,
        "profitability_message": profitability_message,
        "profitability_status": profitability_status,
        
        # Wartości w PLN dla wyświetlania
        "target_sales_pln": target_sales_pln,
        "marketing_budget_pln": marketing_budget_pln,
        "gross_profit_pln": gross_profit_pln,
        "net_profit_pln": net_profit_pln,
        
        # Dodatkowe wskaźniki
        "profit_margin": round((net_profit / target_sales) * 100, 1) if target_sales > 0 else 0,
        "marketing_to_profit_ratio": round((marketing_budget / net_profit) * 100, 1) if net_profit > 0 else 0,
        
        # Informacje o walucie
        "currency": currency,
        "eur_rate": round(eur_rate, 4),
        "currency_info": f"Kurs EUR/PLN: {eur_rate:.4f} (NBP)"
    }

def create_excel_report(results: Dict[str, Any]) -> BytesIO:
    """
    Tworzy profesjonalny raport Excel z wynikami prognoz ACOS.
    
    Args:
        results (Dict[str, Any]): Wyniki obliczeń
    
    Returns:
        BytesIO: Bufor z plikiem Excel
    """
    # Tworzenie nowego skoroszytu
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise ValueError("Nie udało się utworzyć arkusza Excel")
    
    ws.title = "Prognoza ACOS"
    
    # Kolory firmowe
    orange_fill = PatternFill(start_color="f39c12", end_color="e67e22", fill_type="solid")
    light_orange_fill = PatternFill(start_color="fdf2e9", end_color="fdf2e9", fill_type="solid")
    gray_fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type="solid")
    green_fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
    red_fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
    
    # Czcionki
    title_font = Font(name="Arial", size=18, bold=True, color="2c3e50")
    header_font = Font(name="Arial", size=12, bold=True, color="ffffff")
    subheader_font = Font(name="Arial", size=11, bold=True, color="2c3e50")
    normal_font = Font(name="Arial", size=10, color="2c3e50")
    
    # Wyrównania
    center_alignment = Alignment(horizontal="center", vertical="center")
    left_alignment = Alignment(horizontal="left", vertical="center")
    right_alignment = Alignment(horizontal="right", vertical="center")
    
    # Ramki
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # Nagłówek raportu
    ws.merge_cells("A1:J1")
    ws["A1"] = "🎯 PROGNOZA ACOS - RAPORT ANALITYCZNY"
    ws["A1"].font = title_font
    ws["A1"].alignment = center_alignment
    ws["A1"].fill = light_orange_fill
    
    # Informacje o raporcie
    ws.merge_cells("A2:J2")
    ws["A2"] = f"Wygenerowano: {datetime.now().strftime('%d.%m.%Y %H:%M')} | AmzTeam.pro"
    ws["A2"].font = Font(name="Arial", size=10, italic=True, color="7f8c8d")
    ws["A2"].alignment = center_alignment
    
    # Sekcja 1: Parametry wejściowe
    current_row = 4
    ws.merge_cells(f"A{current_row}:J{current_row}")
    ws[f"A{current_row}"] = "📊 PARAMETRY WEJŚCIOWE"
    ws[f"A{current_row}"].font = subheader_font
    ws[f"A{current_row}"].fill = orange_fill
    ws[f"A{current_row}"].font = Font(name="Arial", size=12, bold=True, color="ffffff")
    ws[f"A{current_row}"].alignment = center_alignment
    
    current_row += 1
    input_params = [
        ("Marża brutto", f"{results.get('gross_margin', 0):.1f}%"),
        ("Docelowe AOV", f"€{results.get('target_aov', 0):.0f} ({results.get('target_aov_pln', 0):.0f} PLN)"),
        ("Docelowy CTR", f"{results.get('target_ctr', 0):.2f}%"),
        ("Docelowy CPC", f"€{results.get('target_cpc', 0):.2f} ({results.get('target_cpc_pln', 0):.2f} PLN)"),
        ("Docelowy CVR", f"{results.get('target_cvr', 0):.2f}%"),
        ("Wyświetlenia", f"{results.get('impressions', 0):,}".replace(',', ' '))
    ]
    
    for i, (param, value) in enumerate(input_params):
        row = current_row + i
        ws[f"A{row}"] = param
        ws[f"A{row}"].font = normal_font
        ws[f"A{row}"].alignment = left_alignment
        ws[f"B{row}"] = value
        ws[f"B{row}"].font = Font(name="Arial", size=10, bold=True, color="2c3e50")
        ws[f"B{row}"].alignment = right_alignment
        
        # Dodanie obramowania
        for col in ["A", "B"]:
            ws[f"{col}{row}"].border = thin_border
            if i % 2 == 0:
                ws[f"{col}{row}"].fill = gray_fill
    
    # Sekcja 2: Wyniki prognozy
    current_row += len(input_params) + 2
    ws.merge_cells(f"A{current_row}:J{current_row}")
    ws[f"A{current_row}"] = "🚀 WYNIKI PROGNOZY"
    ws[f"A{current_row}"].font = Font(name="Arial", size=12, bold=True, color="ffffff")
    ws[f"A{current_row}"].fill = orange_fill
    ws[f"A{current_row}"].alignment = center_alignment
    
    current_row += 1
    forecast_results = [
        ("Prognozowana sprzedaż", f"€{results.get('projected_sales', 0):,}".replace(',', ' '), f"{results.get('projected_sales_pln', 0):,} PLN".replace(',', ' ')),
        ("Prognozowane wydatki", f"€{results.get('projected_spend', 0):,}".replace(',', ' '), f"{results.get('projected_spend_pln', 0):,} PLN".replace(',', ' ')),
        ("Oczekiwany ACOS", f"{results.get('acos', 0):.0f}%", ""),
        ("ROI", f"{results.get('roi', 0):.0f}%", ""),
        ("Zysk na sprzedaż", f"€{results.get('profit_per_sale', 0):,}".replace(',', ' '), f"{results.get('profit_per_sale_pln', 0):,} PLN".replace(',', ' ')),
        ("Całkowity zysk", f"€{results.get('profit', 0):,}".replace(',', ' '), f"{results.get('profit_pln', 0):,} PLN".replace(',', ' ')),
        ("Prognozowane kliknięcia", f"{results.get('clicks', 0):,}".replace(',', ' '), ""),
        ("Prognozowane zamówienia", f"{results.get('orders', 0):,}".replace(',', ' '), ""),
        ("Break-even ACOS", f"{results.get('break_even_acos', 0):.0f}%", ""),
        ("ROAS", f"{results.get('roas', 0):.2f}", ""),
        ("CPM", f"€{results.get('cpm', 0):.2f}", ""),
        ("Koszt na konwersję", f"€{results.get('cost_per_conversion', 0):.2f}", "")
    ]
    
    # Nagłówki kolumn
    ws[f"A{current_row}"] = "WSKAŹNIK"
    ws[f"B{current_row}"] = "WARTOŚĆ EUR"
    ws[f"C{current_row}"] = "WARTOŚĆ PLN"
    
    for col in ["A", "B", "C"]:
        ws[f"{col}{current_row}"].font = header_font
        ws[f"{col}{current_row}"].fill = orange_fill
        ws[f"{col}{current_row}"].alignment = center_alignment
        ws[f"{col}{current_row}"].border = thin_border
    
    current_row += 1
    for i, (metric, eur_value, pln_value) in enumerate(forecast_results):
        row = current_row + i
        ws[f"A{row}"] = metric
        ws[f"A{row}"].font = normal_font
        ws[f"A{row}"].alignment = left_alignment
        
        ws[f"B{row}"] = eur_value
        ws[f"B{row}"].font = Font(name="Arial", size=10, bold=True, color="2c3e50")
        ws[f"B{row}"].alignment = right_alignment
        
        ws[f"C{row}"] = pln_value
        ws[f"C{row}"].font = normal_font
        ws[f"C{row}"].alignment = right_alignment
        
        # Kolorowanie wierszy
        fill_color = gray_fill if i % 2 == 0 else PatternFill(start_color="ffffff", end_color="ffffff", fill_type="solid")
        
        # Specjalne kolorowanie dla rentowności
        if "zysk" in metric.lower() or "profit" in metric.lower():
            profit_value = results.get('profit', 0)
            if profit_value > 0:
                fill_color = green_fill
            elif profit_value < 0:
                fill_color = red_fill
        
        for col in ["A", "B", "C"]:
            ws[f"{col}{row}"].border = thin_border
            ws[f"{col}{row}"].fill = fill_color
    
    # Sekcja 3: Analiza rentowności
    current_row += len(forecast_results) + 2
    ws.merge_cells(f"A{current_row}:J{current_row}")
    ws[f"A{current_row}"] = "💡 ANALIZA RENTOWNOŚCI"
    ws[f"A{current_row}"].font = Font(name="Arial", size=12, bold=True, color="ffffff")
    ws[f"A{current_row}"].fill = orange_fill
    ws[f"A{current_row}"].alignment = center_alignment
    
    current_row += 1
    
    # Status rentowności
    is_profitable = results.get('is_profitable', False)
    status_text = "✅ KAMPANIA RENTOWNA" if is_profitable else "⚠️ KAMPANIA NIERENTOWNA"
    status_color = green_fill if is_profitable else red_fill
    
    ws.merge_cells(f"A{current_row}:C{current_row}")
    ws[f"A{current_row}"] = status_text
    ws[f"A{current_row}"].font = Font(name="Arial", size=11, bold=True, color="2c3e50")
    ws[f"A{current_row}"].fill = status_color
    ws[f"A{current_row}"].alignment = center_alignment
    ws[f"A{current_row}"].border = thin_border
    
    current_row += 1
    
    # Komunikat o rentowności
    profitability_message = results.get('profitability_message', '')
    if profitability_message:
        ws.merge_cells(f"A{current_row}:J{current_row}")
        ws[f"A{current_row}"] = profitability_message.replace('⚠️', '').replace('✅', '').strip()
        ws[f"A{current_row}"].font = normal_font
        ws[f"A{current_row}"].alignment = left_alignment
        ws[f"A{current_row}"].fill = light_orange_fill
        ws[f"A{current_row}"].border = thin_border
    
    # Sekcja 4: Informacje o walucie
    current_row += 3
    ws.merge_cells(f"A{current_row}:J{current_row}")
    ws[f"A{current_row}"] = "💱 INFORMACJE O WALUCIE"
    ws[f"A{current_row}"].font = Font(name="Arial", size=12, bold=True, color="ffffff")
    ws[f"A{current_row}"].fill = orange_fill
    ws[f"A{current_row}"].alignment = center_alignment
    
    current_row += 1
    currency_info = results.get('currency_info', '')
    ws.merge_cells(f"A{current_row}:J{current_row}")
    ws[f"A{current_row}"] = currency_info
    ws[f"A{current_row}"].font = normal_font
    ws[f"A{current_row}"].alignment = center_alignment
    ws[f"A{current_row}"].fill = light_orange_fill
    ws[f"A{current_row}"].border = thin_border
    
    # Stopka
    current_row += 3
    ws.merge_cells(f"A{current_row}:J{current_row}")
    ws[f"A{current_row}"] = "🔧 Wygenerowano przez AmzTeam.pro | Kalkulator ACOS"
    ws[f"A{current_row}"].font = Font(name="Arial", size=9, italic=True, color="7f8c8d")
    ws[f"A{current_row}"].alignment = center_alignment
    
    current_row += 1
    ws.merge_cells(f"A{current_row}:J{current_row}")
    ws[f"A{current_row}"] = "📧 Kontakt: Bartek z Twoje Drzwi do Amazon | Powered by Cursor AI"
    ws[f"A{current_row}"].font = Font(name="Arial", size=9, italic=True, color="7f8c8d")
    ws[f"A{current_row}"].alignment = center_alignment
    
    # Dostosowanie szerokości kolumn
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    
    # Zapisanie do bufora
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer

def format_currency(amount: float, currency: str = "EUR") -> str:
    """
    Formatuje kwotę w podanej walucie.
    
    Args:
        amount (float): Kwota do sformatowania
        currency (str): Kod waluty (EUR, PLN)
    
    Returns:
        str: Sformatowana kwota
    """
    if currency == "EUR":
        return f"€{amount:,.0f}".replace(',', ' ')
    elif currency == "PLN":
        return f"{amount:,.2f} PLN".replace(',', ' ')
    else:
        return f"{amount:,.2f} {currency}".replace(',', ' ')

def format_percentage(value: float) -> str:
    """
    Formatuje wartość procentową.
    
    Args:
        value (float): Wartość do sformatowania
    
    Returns:
        str: Sformatowana wartość procentowa
    """
    return f"{value:.2f}%"

def validate_input_data(sales: float, spend: float, margin: float) -> Dict[str, Any]:
    """
    Waliduje dane wejściowe do obliczeń.
    
    Args:
        sales (float): Prognozowana sprzedaż
        spend (float): Prognozowane wydatki na reklamę
        margin (float): Marża brutto w procentach
    
    Returns:
        Dict[str, Any]: Słownik z wynikiem walidacji
    """
    errors = []
    
    if sales < 0:
        errors.append("Prognozowana sprzedaż musi być dodatnia")
    
    if spend < 0:
        errors.append("Prognozowane wydatki muszą być dodatnie")
    
    if margin < 0:
        errors.append("Marża brutto musi być dodatnia")
    
    if margin > 100:
        errors.append("Marża brutto nie może przekraczać 100%")
    
    if sales > 0 and spend > sales:
        errors.append("Wydatki na reklamę nie mogą być większe niż prognozowana sprzedaż")
    
    return {
        "is_valid": len(errors) == 0,
        "errors": errors
    }

def generate_export_data(results: Dict[str, Any]) -> Dict[str, Any]:
    """
    Generuje dane do eksportu w formacie odpowiednim do pobrania.
    
    Args:
        results (Dict[str, Any]): Wyniki obliczeń ACOS
    
    Returns:
        Dict[str, Any]: Dane do eksportu
    """
    export_data = {
        "export_timestamp": datetime.now().isoformat(),
        "forecast_results": {
            "projected_sales_eur": results.get("projected_sales", 0),
            "projected_spend_eur": results.get("projected_spend", 0),
            "expected_acos_percent": results.get("acos", 0),
            "roi_percent": results.get("roi", 0),
            "profit_per_sale_eur": results.get("profit_per_sale", 0),
            "total_profit_eur": results.get("profit", 0),
            "break_even_acos_percent": results.get("break_even_acos", 0),
            "projected_clicks": results.get("clicks", 0),
            "projected_orders": results.get("orders", 0),
            "roas": results.get("roas", 0),
            "cpm_eur": results.get("cpm", 0),
            "cost_per_conversion_eur": results.get("cost_per_conversion", 0)
        },
        "input_parameters": {
            "gross_margin_percent": results.get("gross_margin", 0),
            "target_aov_eur": results.get("target_aov", 0),
            "target_ctr_percent": results.get("target_ctr", 0),
            "target_cpc_eur": results.get("target_cpc", 0),
            "target_cvr_percent": results.get("target_cvr", 0),
            "impressions": results.get("impressions", 0)
        },
        "profitability_analysis": {
            "is_profitable": results.get("is_profitable", False),
            "profitability_status": results.get("profitability_status", "unknown"),
            "profitability_message": results.get("profitability_message", "")
        },
        "currency_info": {
            "primary_currency": "EUR",
            "eur_pln_rate": results.get("eur_rate", 4.30),
            "currency_source": "NBP API"
        },
        "acos_formula": "ACOS = Ad Spend / Ad Sales = (Bid × Clicks) / (Orders × AOV)",
        "recommendations": {
            "target_acos_max": results.get("break_even_acos", 0),
            "current_performance": "profitable" if results.get("is_profitable", False) else "unprofitable"
        }
    }
    
    return export_data 